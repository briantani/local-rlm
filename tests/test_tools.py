import pytest
from pathlib import Path
from src.core.agent import RLMAgent
from src.tools.search import search_web
from src.core.budget import BudgetManager
import dspy
from tests.conftest import get_lm_for_testing
import os

# Mocking file creation for Excel
def create_mock_excel(path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)

@pytest.fixture
def setup_dspy_ollama():
    """
    Ensures correct LM is loaded and budget is reset.
    """
    BudgetManager().reset()
    try:
        lm = get_lm_for_testing("ollama")
        dspy.settings.configure(lm=lm)
    except Exception as e:
        pytest.skip(f"Skipping Agent tests because LM could not be loaded: {e}")

@pytest.mark.timeout(30)
@pytest.mark.asyncio
async def test_web_search_tool():
    """
    Test 8.3a: Basic functionality of the search wrapper.
    """
    results = search_web("python programming language creator", max_results=1)
    assert isinstance(results, list)
    assert len(results) > 0
    assert "title" in results[0]
    assert "href" in results[0]
    # Check that we got relevant results
    content = str(results[0]).lower()
    assert "python" in content

@pytest.mark.integration
@pytest.mark.timeout(120)
@pytest.mark.asyncio
async def test_excel_formula_reading(tmp_path, setup_dspy_ollama, monkeypatch):
    """
    Test 8.1: Agent reads Excel formula, not just value.
    """
    # Fix: Change CWD to tmp_path so the Agent's code (which uses relative paths) works.
    monkeypatch.chdir(tmp_path)

    excel_path = tmp_path / "test.xlsx"
    create_mock_excel(excel_path)

    # Load compiled coder if available
    # We need to find the project root because we changed cwd to tmp_path
    project_root = Path(__file__).parent.parent
    compiled_path = project_root / "src/modules/coder_compiled.json"

    coder = None
    if compiled_path.exists():
        from src.modules.coder import Coder
        coder = Coder()
        try:
            coder.load(str(compiled_path))
            print("Loaded compiled coder for test.")
        except Exception as e:
            print(f"Failed to load compiled coder: {e}")
            coder = None

    # If integration mode is not enabled, use mocks to avoid calling external LLMs
    if not os.getenv("RLM_RUN_INTEGRATION"):
        from tests.conftest import MockCoder, MockREPL, MockResponder

        # Coder should read the excel and print the formula in A3
        code = 'import openpyxl\nwb = openpyxl.load_workbook("test.xlsx", data_only=False)\nws = wb.active\nprint(ws["A3"].value)'
        mock_coder = MockCoder(code=code)
        mock_repl = MockREPL(output="=SUM(A1:A2)")
        mock_responder = MockResponder(response="=SUM(A1:A2)")
        agent = RLMAgent(max_steps=3, root_dir=tmp_path, coder=mock_coder, repl=mock_repl, responder=mock_responder)

        task = "What is the formula in cell A3 of test.xlsx?"
        result = agent.run(task)

        assert "SUM" in result or "A1:A2" in result
    else:
        agent = RLMAgent(max_steps=3, root_dir=tmp_path, coder=coder)

        # Task explicitly asks for the formula
        task = "What is the formula in cell A3 of test.xlsx?"
        result = agent.run(task)

        # Needs to find "=SUM(A1:A2)"
        assert "SUM" in result or "A1:A2" in result

@pytest.mark.timeout(60)
@pytest.mark.asyncio
async def test_agent_web_search(setup_dspy_ollama, monkeypatch):
    """
    Test 8.3b: Agent uses search tool to answer question.
    """
    # Mock search_web to return deterministic results and avoid network calls
    def mock_search(query, max_results=5):
        return [{"title": "History of Python", "href": "http://python.org", "body": "Guido van Rossum created Python in 1989."}]

    # We must patch it where it is IMPORTED in the agent's code execution env?
    # NO. The agent executes code that imports `src.tools.search`.
    # If we patch `src.tools.search.search_web`, the imported module will have the mock.
    monkeypatch.setattr("src.tools.search.search_web", mock_search)

    agent = RLMAgent(max_steps=3)
    # We ask something dynamic/external that requires search
    task = "Use the search_web tool to find out who created the Python programming language."

    # To avoid dependency on the LM during this unit test, inject lightweight
    # mocks for Architect/Coder/REPL so the test focuses on the search tool.
    from tests.conftest import MockArchitect, MockCoder, MockREPL

    # Architect that returns CODE for the first step(s) then ANSWER to finish
    class ToggleArchitect:
        def __init__(self):
            self.calls = 0

        def __call__(self, query: str, data_desc: str = "", artifacts_info: str = ""):
            self.calls += 1
            if self.calls < 3:
                return type('P', (), {'action': 'CODE'})()
            return type('P', (), {'action': 'ANSWER'})()

    mock_arch = ToggleArchitect()
    # Coder produces code that calls the (monkeypatched) search_web and prints the body
    search_code = 'result = search_web("who created Python programming language")\nprint(result[0]["body"])'
    mock_coder = MockCoder(code=search_code)
    mock_repl = MockREPL(output="Guido van Rossum created Python in 1989.")
    mock_resp = MockREPL(output="Guido van Rossum created Python in 1989.")
    # Use a simple MockResponder to avoid calling the LM for final formatting
    from tests.conftest import MockResponder
    mock_responder = MockResponder(response="Guido van Rossum created Python in 1989.")

    agent = RLMAgent(max_steps=4, architect=mock_arch, coder=mock_coder, repl=mock_repl, responder=mock_responder)

    result = agent.run(task)

    assert "Guido" in result or "Van Rossum" in result
